 from openpyxl import load_workbook
#from openpyxl.styles.colors import Color


if __name__ == '__main__':
	print "World parser initiated!"
	wb = load_workbook('Chart.xlsx')
	ws = wb["Sheet1"]


	cell_range = "A1:B2"
	for row in ws.iter_rows(cell_range):
		#print row[0].value
		for cell in row:
			print cell.fill.bgColor.rgb


def parse_color_chart(chart_worksheet):
	color_dict = {}
	
	cell_range = A1:B2
	for row in chart_worksheet.iter_rows(cell_range):
		print row[0].value
		color_dict[row[0].value] = row[1].value


	return color_dict

def parse_world():
	pass

def save_world():
	pass
